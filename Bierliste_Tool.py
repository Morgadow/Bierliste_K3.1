#!/usr/bin/python3.7
# -*- coding: utf-8 -*-

import os
import sys
import datetime
import shutil
import lib.logger as logging
import math
from lib.install_dep import DEPENDECIES, install_dep
try:
    import tkinter as tk
    from tkinter import messagebox
    import openpyxl as opxl
    from configparser import ConfigParser
except Exception as e:
    print(e)
    print("Missing dependencies, proceeding to install {} packages:".format(len(DEPENDECIES)))
    install_dep()

# todo format ausgabe excel tabellen
# todo Stand_XX do not print nulls
# todo fenster person ändern (zimmer, name, ...)
# todo fenster schöner machen, hintergrund, buttons, ...
# todo pillow resize image und childs mit background
# todo exe mit data
# todo excel example in container --> generate default
# todo küche aus settings in Titel excel string

# version
__major__ = 0  # for major interface/format changes
__minor__ = 0  # for minor interface/format changes
__release__ = 2  # for tweaks, bug-fixes, or development
__version__ = '%d.%d.%d' % (__major__, __minor__, __release__)
__version_info__ = tuple([int(num) for num in __version__.split('.')])
__author__ = "Simon Schmid"
__date__ = '13.01.2020'

# globals
LOG_LEVEL = logging.DEBUG
SETTINGS_FILE = "settings.ini"
EXPORT_FOLDER = 'Bierlisten'
EXAMPLE_EXCEL = 'Example_file.xlsx'
HELP_FILE = 'Anleitung.pdf'
SMALL_LABEL_FONT = "Helvetica 9 bold"

EXCEL_START_ROW = 3
STD_VALUES = {'room': '', 'balance': 0.0, 'beers': 0, 'radler': 0, 'mate': 0, 'pali': 0, 'spezi': 0}
STD_COLS = {'room': 'A', 'name': 'B', 'balance': 'C',  'new_beer': 'D', 'new_radler': 'E', 'new_mate': 'F', 'new_pali': 'G', 'new_spezi': 'H',  'beers': 'I', 'radler': 'J', 'mate': 'K', 'pali': 'L', 'spezi': 'M'}
ROOMS_OWN_KITCHEN = ('310', '311', '312', '313', '314', '315', '316', '317', '318', '319', '349')

# size of main self.root (optimal sizes for chosen background image)
HEIGHT = 450
WIDTH = 600


def handle_excep(exception, with_tb=True):
    """ prints exception """
    logging.static_critical(exception)
    if with_tb:
        import traceback
        logging.static_critical(traceback.format_exc())


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def select_file():
    try:
        from tkinter import filedialog
        file = os.path.normpath(os.path.abspath(filedialog.askopenfilename(title="Select file", filetypes=(("Excel file", "*.xlsx *.xls *xlsm"), ("Alle Dateien", "*.*")))))
        if file is not None and file != '' and os.path.isfile(file):
            logging.static_info("Selected file: {}".format(file))
            return file
        else:
            logging.static_warning("Keine Datei gewählt!")
            return None
    except Exception as exp:
        handle_excep(exp, with_tb=True)
        return None


def ask_user_yn(message):
    """
    Asks user message 'message' and returns true/y or false/n
    :param message: String, question to ask
    :return: Boolean, response
    """
    try:
        while True:
            raw_in = input(message.strip() + ' (y/n) ')
            if raw_in == 'y':
                logging.static_debug("User response for question '{}': True".format(message.strip))
                return True
            elif raw_in == 'n':
                logging.static_debug("User response for question '{}': False".format(message.strip))
                return False
    except Exception as exp:
        handle_excep(exp)
        return None


# noinspection PyBroadException
class BierListeTool:

    def __init__(self):

        try:  # some builds fails mysteriously

            self.logger = logging.Logger(level=LOG_LEVEL)

            self.prices = SettingsGroup()
            self.prices.beer = None
            self.prices.radler = None
            self.prices.mate = None
            self.prices.pali = None
            self.prices.spezi = None
            self.kueche = None
            self.read_settings_file(SETTINGS_FILE)

            self.today = datetime.datetime.now().strftime('%d.%m.%Y')

            self.drinker = []  # every person

            self._build_GUI()

        except Exception as ex:
            handle_excep(ex)
            logging.static_critical("Could not build GUI from this path, copy to local system and retry!")
            print("CRITICAL: Could not build GUI from this path, start with START.bat or copy to local system and retry!")
            print("Alternatively, start tool via the terminal")
            print("\t1) Open the terminal with 'cmd' in startmenu")
            print("\t2) navigate to scriptpath with command: 'pushd [Scriptpath]'")
            print("\t3) Start tool with command: Bierliste_Tool.py")
            print("\t4) If no Python version is installed please try START.bat")
            print("")
            os.system("pause")

    def read_settings_file(self, settings_file):
        """
        Reads price list from settings .ini file
        :param settings_file: String, Name of settings file, type .ini
        :return: None
        """

        if not os.path.isfile(settings_file):
            raise FileNotFoundError("CRITICAL: Settings file not found: " + str(settings_file))

        # init configparser and get information
        self.logger.debug("Reading settings file:")
        config = ConfigParser()
        config.read(settings_file)

        try:
            # prices
            self.prices.beer = float(config.get('Preise', 'Bier'))
            self.prices.radler = float(config.get('Preise', 'Radler'))
            self.prices.mate = float(config.get('Preise', 'Mate'))
            self.prices.pali = float(config.get('Preise', 'Pali'))
            self.prices.spezi = float(config.get('Preise', 'Spezi'))
            self.prices.add_charge = float(config.get('Preise', 'Aufpreis_Externe'))
            self.kueche = config.get('General', 'Kueche').replace('"', '')

            self.logger.info("Successfully imported pricelist!")
            self.logger.debug("Preise | Bier: {}, Radler: {}, Mate: {}, Pali: {}, Spezi: {}".format(self.prices.beer, self.prices.radler, self.prices.mate, self.prices.pali, self.prices.spezi))
        except Exception as ex:
            self.logger.error("Could not read Settings file!")
            handle_excep(ex, with_tb=True)

    @staticmethod
    def generate_default_settingfile():
        """ generate default ini settingsfile in scriptfolder """
        with open(os.path.join(os.getcwd(), SETTINGS_FILE), 'w+') as ini_file:
            ini_file.write('[Preise]' + '\n')
            ini_file.write('Bier = 1.00' + '\n')
            ini_file.write('Radler = 1.00' + '\n')
            ini_file.write('Mate = 1.00' + '\n')
            ini_file.write('Pali = 1.00' + '\n')
            ini_file.write('Spezi = 1.00' + '\n')
            ini_file.write('Aufpreis_Externe = 0.05' + '\n')
            if os.path.exists(os.path.join(os.getcwd(), SETTINGS_FILE)):
                logging.static_info("Generated default {} file".format(SETTINGS_FILE))
            else:
                logging.static_error("Could not generate default {} file".format(SETTINGS_FILE))

    def _build_GUI(self):
        """ Builds gui and holds mainloop """

        self.logger.debug("Building GUI:")

        self.root = tk.Tk()
        self.root.title("Bierwart Helper")
        self.root.resizable(False, False)
        tk.Canvas(self.root, height=HEIGHT, width=WIDTH).pack()

        # background label for background image and program icon
        try:
            self.root.wm_iconbitmap(bitmap=resource_path("gui\\icon.ico"))
            background_image = tk.PhotoImage(file=resource_path("gui\\background.png"))
            tk.Label(self.root, image=background_image).place(relwidth=1, relheight=1)
        except Exception:
            self.root.wm_iconbitmap(bitmap=resource_path("icon.ico"))
            background_image = tk.PhotoImage(file=resource_path("background.png"))
            tk.Label(self.root, image=background_image).place(relwidth=1, relheight=1)

        import_excel_btn = tk.Button(self.root, bd=5, font=SMALL_LABEL_FONT, bg='gray', text="Import Excel", command=lambda: self._import_excel())
        import_excel_btn.place(relx=0.2125, rely=0.9, relwidth=0.175, relheight=0.075)

        export_excel_btn = tk.Button(self.root, font=SMALL_LABEL_FONT, bd=5, bg='gray', text="Export to Excel", command=lambda: self._export_excel())
        export_excel_btn.place(relx=0.6125, rely=0.9, relwidth=0.175, relheight=0.075)

        new_person_btn = tk.Button(self.root, font=SMALL_LABEL_FONT, bd=5, bg='gray', text="Neuer Drinker", command=lambda: self._new_person())
        new_person_btn.place(relx=0.4125, rely=0.9, relwidth=0.175, relheight=0.075)

        # help and credits
        help_btn = tk.Button(self.root, text="Hilfe", bg='lightgrey', bd=2, command=lambda: self._open_file(HELP_FILE))
        help_btn.place(relx=0.025, rely=0.925, relwidth=0.1, relheight=0.05)
        tk.Label(self.root, font=("Arial", 7), text="Version: {}".format(__version__), bg="lightgrey").place(relx=0.855, rely=0.925, relwidth=0.12, relheight=0.05)

        self.logger.info("Successfully build GUI")
        tk.mainloop()

    def _import_excel(self):
        """
        Choose excel file with file dialog and if needed choose a sheet over child window with buttons
        :return: None
        """

        # choose excel file over filedialog
        excel_file = select_file()
        if excel_file is None:
            self.logger.error("No file selected for import!")
        else:

            # select sheet in excel
            workbook = opxl.load_workbook(filename=excel_file)
            if len(workbook.sheetnames) == 1:
                self._read_excel_file(excel_file, workbook.sheetnames[0])
            elif len(workbook.sheetnames) < 1:  # no sheet, should not happen
                messagebox.showwarning('Fehler', "Keine Tabelle in Excel Datei gefunden!")
                self.logger.error("Empty excel file!")
                return
            else:
                self.logger.debug("Build child window for selecting excel sheet to import:")
                child = tk.Toplevel()
                child.resizable(False, False)
                child.title("Select Excel Sheet")
                if len(workbook.sheetnames) > 2:
                    tk.Canvas(child, height=math.ceil(len(workbook.sheetnames)/2)*25+30, width=380).pack()
                else:
                    tk.Canvas(child, height=45, width=380).pack()
                tk.Label(child, bg='white').place(relwidth=1, relheight=1)
                try:
                    child.wm_iconbitmap(bitmap=resource_path("gui\\icon.ico"))
                except Exception:
                    child.wm_iconbitmap(bitmap=resource_path("icon.ico"))
                tk.Label(child, bg='lightgrey').place(relwidth=1, relheight=1)

                top, spacer, elem_width, elem_height = 10, 10, 175, 25
                sheet_btns = []
                for index, elem in enumerate(workbook.sheetnames):
                    sheet_btns.append(tk.Button(child, text=elem, font='Helvetica 9', bg='gray', bd=2, command=lambda c=index: self._read_excel_file(excel_file, sheet_btns[c].cget('text'), child)))
                    sheet_btns[-1].place(x=spacer+((spacer+elem_width)*(index % 2)), y=top+(spacer+elem_height)*math.floor(index/2), width=elem_width, height=elem_height)

    def _read_excel_file(self, excel_file, excel_sheet, child=None):
        """
        Reads content from selected excel file and selected worksheet, destroys child if given, reloads person buttons in root
        :param excel_file: Path, excel file to read
        :param excel_sheet: String, name of worksheet to read, first if only one in document
        :param child: Tk.Toplevel(), Optional if more than one sheet found, gets destroyed
        :return: None
        """

        extracted_data = []
        self.logger.info("Selected worksheet: {}".format(excel_sheet))
        if child is not None:
            child.destroy()

        # read person data while name cell is not empty
        workbook = opxl.load_workbook(filename=excel_file)
        row_counter = EXCEL_START_ROW
        cells_to_read = ('name', 'room', 'balance', 'beers', 'radler', 'mate', 'pali', 'spezi')
        while workbook[excel_sheet]['B{}'.format(row_counter)].value != '' and workbook[excel_sheet]['B{}'.format(row_counter)].value is not None:
            cells = []
            for elem in cells_to_read:
                if workbook[excel_sheet]['{}{}'.format(STD_COLS[elem], row_counter)].value is not None:
                    cells.append(workbook[excel_sheet]['{}{}'.format(STD_COLS[elem], row_counter)].value)
                else:
                    cells.append(STD_VALUES[elem])

            extracted_data.append(Person(cells[0], room=cells[1], balance=cells[2], beers=cells[3], radler=cells[4], mate=cells[5], pali=cells[6], spezi=cells[7]))
            self.logger.debug("Extracted | " + str(extracted_data[-1]))
            row_counter += 1

        self.logger.info("Extracted data for {} person(s)".format(row_counter-EXCEL_START_ROW))
        self.drinker = extracted_data
        self._update_drinker_btns()

    def _update_drinker_btns(self):
        """ Creates buttons for all persons in root """

        self.logger.debug("Starting to update player buttons:")
        try:  # delete all for recreating
            for person in self.drinker:
                if hasattr(person, 'button'):
                    person.button.destroy()
        except Exception as err:  # throws error when called first time
            handle_excep(err)
            # pass  # first time no buttons  # todo wieder rein?

        spacer, width, height = 0.025, 0.17, 0.05
        for index, person in enumerate(self.drinker):
            person.button = tk.Button(self.root, font="Helvetica 8 bold", text=person.name, command=lambda c=index: self._cb_edit_person(self.drinker[c]))
            if person.updated:
                person.button.config(bg='lightgreen')
            else:
                person.button.config(bg='tomato')

        for index, person in enumerate(self.drinker):
            person.button.place(relx=(index % 5)*(spacer+width)+spacer, rely=int(index/5)*(spacer+height)+spacer, relheight=height, relwidth=width)
        self.logger.info("Updated player buttons")

    def _delete_person(self, drinker):
        """
        Deletes drinker from self.drinker und recreates all person_btns
        :param drinker: Person(), drinker to edit
        :return: None
        """
        try:
            self.logger.debug("Trying to delete drinker:" + drinker.name)
            for index, person in enumerate(self.drinker):
                if person.ID == drinker.ID:
                    del self.drinker[index]
                    self.logger.info("Deleted drinker: " + str(drinker.name))
                    break
        except Exception as e:
            messagebox.showerror('Fehler', "Could not delete drinker:" + str(drinker.name))
            handle_excep(e)

    def _new_person(self):
        """
        Callback for btn "New Person", Builds child window for creating a new person
        :return: None
        """
        # build child window to insert information
        self.logger.debug("Build child window for new person:")
        child = tk.Toplevel()
        child.resizable(False, False)
        canvas = tk.Canvas(child, height=200, width=150)
        canvas.pack()
        child.title("Lege neue Person an")
        try:
            child.wm_iconbitmap(bitmap=resource_path("gui\\icon.ico"))
        except Exception:
            child.wm_iconbitmap(bitmap=resource_path("icon.ico"))
        tk.Label(child, bg='lightgrey').place(relwidth=1, relheight=1)

        # child elements
        top, spacer, elem_height = 0.15, 0.025, 0.1
        tk.Label(child, text='Name', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.1, rely=top+spacer, relheight=elem_height, relwidth=0.8)
        name_entry = tk.Entry(child)
        name_entry.place(relx=0.1, rely=top+spacer+elem_height, relheight=elem_height, relwidth=0.8)

        tk.Label(child, text='Zimmer', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.1, rely=top+spacer*2+elem_height*2, relheight=elem_height, relwidth=0.8)
        room_entry = tk.Entry(child)
        room_entry.place(relx=0.1, rely=top+spacer*2+elem_height*3, relheight=elem_height, relwidth=0.8)

        enter_btn = tk.Button(child, text="Bestätigen", bd=4, font=SMALL_LABEL_FONT, bg="gray", command=lambda: self.__new_person_aux(child, name_entry.get(), room_entry.get()))
        enter_btn.place(relx=(1-0.6)/2, rely=1-spacer*2-elem_height*1.75, relheight=elem_height*1.5, relwidth=0.6)

    def __new_person_aux(self, child, new_name, new_room):
        """
        Callback after finishing creating a new person, realoads person buttons on root window
        :param child: Tk.Toplevel(), child window for new person to close
        :param new_name: String, name of new person
        :param new_room: String, room of new person
        :return: None
        """
        for drinker in self.drinker:
            if drinker.name == new_name and drinker.room == new_room:
                messagebox.showwarning("Bereits vorhanden", "Person ist bereits angelegt!")
                child.destroy()
                return
        if new_name:
            child.destroy()
            self.drinker.append(Person(new_name, new_room))
            self._update_drinker_btns()
            self.logger.info("Added person with name: {}, room: {}".format(new_name, new_room))
        else:
            messagebox.showwarning('Warning', "Name darf nicht leer sein!")

    def _export_excel(self):
        """ Copies example file to export folder and renames it to current date, if already existing, deletes it """

        name_new_excel = 'Bierliste_' + self.kueche + '_' + str(self.today) + '.xlsx'

        # create export folder if not there already
        if not os.path.isdir(EXPORT_FOLDER):
            os.mkdir(EXPORT_FOLDER)
            if os.path.isdir(EXPORT_FOLDER):
                self.logger.debug("Created export folder: " + EXPORT_FOLDER)
            else:
                self.logger.error("Could not create export folder: " + EXPORT_FOLDER)
                return

        # delete file for today if already there
        self.logger.debug("Starting to copy example file for export:")
        new_excel_file = os.path.abspath(os.path.join(EXPORT_FOLDER, name_new_excel))
        if os.path.isfile(new_excel_file):
            try:
                os.remove(new_excel_file)
                if os.path.isfile(new_excel_file):
                    self.logger.error("Could not delete already existing: {}".format(os.path.join(EXPORT_FOLDER, name_new_excel)))
                    return
                else:
                    self.logger.debug("Deleted already existing file: {}".format(name_new_excel))
            except Exception as ex:
                handle_excep(ex)

        # move example file and rename
        try:
            shutil.copy2(EXAMPLE_EXCEL, os.path.join(EXPORT_FOLDER, name_new_excel))
            if os.path.isfile(os.path.join(EXPORT_FOLDER, name_new_excel)):
                self.logger.info("Copied example file to: {}".format(os.path.join(EXPORT_FOLDER, name_new_excel)))
            else:
                self.logger.error("Could not copy example file to: ".format(os.path.join(EXPORT_FOLDER, name_new_excel)))
                return
        except Exception as ex:
            handle_excep(ex)

        self._fill_excel_file(new_excel_file)

    def _fill_excel_file(self, excel_file):
        """
        Fills new excel file with two tables for a current overview and for a new table to print
        :param excel_file: Path, new excel file
        :return: None
        """

        workbook = opxl.load_workbook(filename=excel_file)
        workbook.copy_worksheet(workbook.active)  # copy example sheet
        if len(workbook.sheetnames) != 2:
            self.logger.error("Example file does not have two sheets, can not be used as standard file!")
            return
        workbook[workbook.sheetnames[0]].title = "Stand " + str(self.today)
        workbook[workbook.sheetnames[1]].title = "Neue Tabelle " + str(self.today)
        workbook[workbook.sheetnames[0]]['M1'].value = self.today
        workbook[workbook.sheetnames[1]]['M1'].value = self.today

        # fill first sheet with current state
        counter = EXCEL_START_ROW
        for person in self.drinker:
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['room'], counter)].value = person.room
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['name'], counter)].value = person.name
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['balance'], counter)].value = person.balance
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['new_beer'], counter)].value = person.new_beer
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['new_radler'], counter)].value = person.new_radler
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['new_mate'], counter)].value = person.new_mate
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['new_pali'], counter)].value = person.new_pali
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['new_spezi'], counter)].value = person.new_spezi
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['beers'], counter)].value = person.beers
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['radler'], counter)].value = person.radler
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['mate'], counter)].value = person.mate
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['pali'], counter)].value = person.pali
            workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['spezi'], counter)].value = person.spezi
            counter += 1

        # # styles, color every second row
        # print(opxl.styles.colors.COLOR_INDEX)
        # for i in range(EXCEL_START_ROW + 1, len(self.drinker) + EXCEL_START_ROW, 2):
        #     print(i)
        #     for elem in STD_COLS:
        #         workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS[elem], i)].font(color=opxl.styles.colors.COLOR_INDEX[3])

        # centering everything except name and room
        # for i in range(EXCEL_START_ROW, len(self.drinker) + EXCEL_START_ROW):
        #     for elem in STD_COLS:
        #         if elem != 'name' and elem != 'room':
        #             workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS[elem], i)].style.alignment.horizontal = 'center'
        #         elif elem == 'name':
        #             workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS[elem], i)].style.alignment.horizontal = 'left'
        #         elif elem == 'room':
        #             workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS[elem], i)].style.alignment.horizontal = 'right'
        #     workbook[workbook.sheetnames[0]]['{}{}'.format(STD_COLS['balance'], i)].style.openpyxl.styles.Font().bold = True

        # fill second sheet with new table to print
        counter = EXCEL_START_ROW
        for person in self.drinker:
            data = {'room': person.room, 'name': person.name, 'balance': person.balance, 'beers': person.beers + person.new_beer, 'radler': person.radler + person.new_radler, 'mate': person.mate + person.new_mate, 'pali': person.pali + person.new_pali, 'spezi': person.spezi + person.new_spezi}
            for elem in data:
                if data[elem]:
                    workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS[elem], counter)].value = data[elem]
            counter += 1

            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['room'], counter)].value = person.room
            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['name'], counter)].value = person.name
            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['balance'], counter)].value = person.balance
            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['beers'], counter)].value = person.beers + person.new_beer
            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['radler'], counter)].value = person.radler + person.new_radler
            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['mate'], counter)].value = person.mate + person.new_mate
            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['pali'], counter)].value = person.pali + person.new_pali
            # workbook[workbook.sheetnames[1]]['{}{}'.format(STD_COLS['spezi'], counter)].value = person.spezi + person.new_spezi
            # counter += 1

        try:
            workbook.save(excel_file)
            self.logger.info("Saved excel file: " + os.path.basename(excel_file))
        except Exception as ex:
            handle_excep(ex)
            self.logger.error("Could not save file, maybe opened!")
            from tkinter import messagebox
            tk.messagebox.showerror('Fehler, Zugriff verweigert!', "Excel Datei konnte nicht gespeichert werden, ist vielleicht geöffnet?!")

    def _person_by_name(self, name):
        """
        Returns list index of person by name
        :param name: String, name of person, same as in excel file
        :return: Integer, index in list self.drinker
        """
        for index, person in enumerate(self.drinker):
            if self.drinker[index].name == name:
                return index, person
        return None

    def _person_by_ID(self, drinker_id):
        """
        Returns list index of person by name
        :param drinker_id: Integer, id of Person() instance
        :return: Integer, index in list self.drinker
        """
        for index, person in enumerate(self.drinker):
            if self.drinker[index].ID == drinker_id:
                return index, person
        return None

    def _cb_edit_person(self, drinker):
        """
        Callback for person buttons, Creates child window for editing data for a person
        :param drinker: Person(), drinker to edit
        :return: None
        """

        # build child window to insert information
        self.logger.debug("Build child window for editing drinker data for: {} in room {}".format(drinker.name, drinker.room))
        child = tk.Toplevel()
        child.title("{} - Zimmer: {}".format(drinker.name, drinker.room))
        child.resizable(False, False)
        tk.Canvas(child, height=250, width=175).pack()
        try:
            child.wm_iconbitmap(bitmap=resource_path("gui\\icon.ico"))
        except Exception:
            child.wm_iconbitmap(bitmap=resource_path("icon.ico"))
        tk.Label(child, bg='lightgrey').place(relwidth=1, relheight=1)

        # child elements
        top, spacer, heigth_elem = 0.1, 0.025, 0.05
        tk.Label(child, anchor='c', text='{} - {}'.format(drinker.name, drinker.room), bg='lightgrey', fg='gray21', font="Helvetica 9 bold italic").place(relx=0.1, rely=spacer, relwidth=0.8, relheight=heigth_elem)
        tk.Label(child, anchor='c', text='Eingezahlt', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.3, rely=spacer+top, relwidth=0.4, relheight=heigth_elem)
        balance_entry = tk.Entry(child)
        balance_entry.place(relx=0.3, rely=spacer+heigth_elem+top, relwidth=0.4, relheight=heigth_elem)

        tk.Label(child, anchor='c', text='Bier', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.4, rely=spacer*2+heigth_elem*2+top, relwidth=0.2, relheight=heigth_elem)
        beer_entry = tk.Entry(child)
        beer_entry.place(relx=0.3, rely=spacer*2+heigth_elem*3+top, relwidth=0.4, relheight=heigth_elem)

        tk.Label(child, anchor='c', text='Radler', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.35, rely=spacer*3+heigth_elem*4+top, relwidth=0.3, relheight=heigth_elem)
        radler_entry = tk.Entry(child)
        radler_entry.place(relx=0.3, rely=spacer*3+heigth_elem*5+top, relwidth=0.4, relheight=heigth_elem)

        tk.Label(child, anchor='c', text='Mate', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.4, rely=spacer*4+heigth_elem*6+top, relwidth=0.2, relheight=heigth_elem)
        mate_entry = tk.Entry(child)
        mate_entry.place(relx=0.3, rely=spacer*4+heigth_elem*7+top, relwidth=0.4, relheight=heigth_elem)

        tk.Label(child, anchor='c', text='Pali', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.4, rely=spacer*5+heigth_elem*8+top, relwidth=0.2, relheight=heigth_elem)
        pali_entry = tk.Entry(child)
        pali_entry.place(relx=0.3, rely=spacer*5+heigth_elem*9+top, relwidth=0.4, relheight=heigth_elem)

        tk.Label(child, anchor='c', text='Spezi', bg='lightgrey', font=SMALL_LABEL_FONT).place(relx=0.4, rely=spacer*6+heigth_elem*10+top, relwidth=0.2, relheight=heigth_elem)
        spezi_entry = tk.Entry(child)
        spezi_entry.place(relx=0.3, rely=spacer*6+heigth_elem*11+top, relwidth=0.4, relheight=heigth_elem)

        enter_btn = tk.Button(child, text="Bestätigen", anchor='c', font=SMALL_LABEL_FONT, bd=4, bg='gray', command=lambda: [self._update_person(drinker, balance_entry.get(), beer_entry.get(), radler_entry.get(), mate_entry.get(), pali_entry.get(), spezi_entry.get()), child.destroy()])
        enter_btn.place(relx=spacer*2, rely=1-spacer-heigth_elem*1.5, relwidth=0.45, relheight=heigth_elem*1.5)
        delete_person_btn = tk.Button(child, text='Löschen', anchor='c', font=SMALL_LABEL_FONT, bd=4, bg='gray', command=lambda: [self._delete_person(drinker.ID), child.destroy()])
        delete_person_btn.place(relx=1-spacer-0.45, rely=1-spacer-heigth_elem*1.5, relwidth=0.45, relheight=heigth_elem*1.5)

    def _update_person(self, drinker, amount, beers, radler, mate, pali, spezi):
        """
        Callback after finishing updating person, Updates data for a person
        :param drinker: Person(), drinker to edit
        :param amount: Float, payed amount of money
        :param beers: Integer, new beers
        :param radler: Integer, new radler
        :param mate: Integer, new mate
        :param pali: Integer, new pali
        :param spezi: Integer, new spezi
        :return: None
        """
        try:
            # handle empty inputs
            make_int = {'beers': beers, 'radler': radler, 'mate': mate, 'pali': pali, 'spezi': spezi}
            for elem in make_int:
                if not make_int[elem]:
                    make_int[elem] = 0
                else:
                    make_int[elem] = int(make_int[elem])

            if amount == '':
                amount = 0
            else:
                amount = round(float(amount.replace(',', '.')), 2)  # can only convert dot type float

            if amount:
                drinker.add_money(amount=amount)
            drinker.add_drinks(beers=make_int['beers'], radler=make_int['radler'], mate=make_int['mate'], pali=make_int['pali'], spezi=make_int['spezi'])
            drinker.bill_drinks(prices=self.prices, beers=make_int['beers'], radler=make_int['radler'], mate=make_int['mate'], pali=make_int['pali'], spezi=make_int['spezi'])
            drinker.updated = True
            drinker.button.config(bg='lightgreen')
        except Exception as err:
            handle_excep(err)
            messagebox.showerror('Fehler', "Could not update Person Information! Check Input")

    def _open_file(self, file):
        """
        Opens file from container or folder
        :param file: Path, file to open with default program
        :return: None
        """
        try:
            file = resource_path(file)
            if os.path.isfile(file):
                import subprocess
                subprocess.Popen([file], shell=True)
            else:
                self.logger.error("Could not find file!: {}".format(file))
        except Exception as err:
            handle_excep(err)


class Person:
    """ Represents one person, holding all corresponding information """

    def __init__(self, name, room='', balance=0, beers=0, radler=0, mate=0, pali=0, spezi=0):
        """

        :param name: String, name
        :param room: String, roomnumber
        :param balance: Float, money account
        :param beers: Integer, number of beers
        :param radler: Integer, number of radler
        :param mate: Integer, number of mate
        :param pali: Integer, number of pali
        :param spezi: Integer, number of spezi
        """
        self.updated = False
        self.logger = logging.Logger()

        self.ID = id(self)
        self.name = name.strip()
        self.room = str(room).strip()
        # self.id = self.generate_id()  # some obsolete approach, use id() instead
        self.balance = balance
        self.beers = beers
        self.radler = radler
        self.mate = mate
        self.pali = pali
        self.spezi = spezi
        self.new_beer = 0
        self.new_radler = 0
        self.new_mate = 0
        self.new_pali = 0
        self.new_spezi = 0

    @staticmethod
    def new_person(name, room, balance=0, beers=0, radler=0, mate=0, pali=0, spezi=0):
        """
        Creates new instance of person
        :param name: String, name
        :param room: String, room
        :param balance: Float, money account
        :param beers: Integer, number of beers
        :param radler: Integer, number of radler
        :param mate: Integer, number of mate
        :param pali: Integer, number of pali
        :param spezi: Integer, number of spezi
        """
        return Person(name.strip(), room=str(room).strip(), balance=round(balance, 2), beers=beers, radler=radler, mate=mate, pali=pali, spezi=spezi)

    def bill_drinks(self, prices, beers, radler, mate, pali, spezi):
        """ Reduces balance by cost of given drinks """
        amount = 0.0
        amount += prices.beer*beers
        amount += prices.radler*radler
        amount += prices.mate*mate
        amount += prices.pali*pali
        amount += prices.spezi*spezi
        self.balance -= round(amount, 2)
        self.logger.info('{} | Billed for drinks:  {} Euro. New balance: {} Euro'.format(self.name, amount, round(self.balance, 2)))
        if str(self.room) not in ROOMS_OWN_KITCHEN:
            self.balance -= round((beers + radler + mate + pali + spezi)*prices.add_charge, 2)
            self.logger.info('{} | Additional extern charge of {} for {} drinks: {} Euro to new balance of {} Euro'.format(self.name, prices.add_charge, beers + radler + mate + pali + spezi, float(round((beers + radler + mate + pali + spezi)*prices.add_charge, 2)), round(self.balance, 2)))

    def add_drinks(self, beers, radler, mate, pali, spezi):
        """ Adds drinks to person """
        self.new_beer = int(beers)
        self.new_radler = int(radler)
        self.new_mate = int(mate)
        self.new_pali = int(pali)
        self.new_spezi = int(spezi)
        self.logger.info("{} | Added {} Bier, {} Radler {} Mate, {} Pali, {} Spezi".format(self.name, beers, radler, mate, pali, spezi))

    def add_money(self, amount):
        """ adds money to balance of user """
        self.balance += round(amount, 2)
        self.logger.info("{} | Added {} Euro to {} Euro to new amount of {} Euro".format(self.name, round(amount, 2), round(self.balance-amount, 2), round(self.balance, 2)))

    def change_room(self, new_room):
        """
        Changes room of person
        :param new_room: String, new room
        :return: None
        """
        if str(new_room) != '':
            self.logger.info("{} | Changed room from {} to {}".format(self.name, self.room, new_room))
            self.room = new_room

    def generate_id(self):
        """
        Generates a ascii based ID for an instance of Person() from name and room
        :return: Integer, unique id of a person
        """
        identifier = []
        identifier.extend([str(ord(elem)) for elem in self.name])
        identifier.extend([str(ord(elem)) for elem in str(self.room)])
        return int(''.join(identifier))

    def __str__(self):
        return "Person | ID: {}, Name: {}, Room: {}, balance: {} Euro, Bier: {}, Radler: {}, Mate: {}, Pali: {}, Spezi: {}".format(self.ID, self.name, self.room, round(self.balance, 2), self.beers, self.radler, self.mate, self.pali, self.spezi)

    def __repr__(self):
        return '\n' + self.__str__() + '\n'


class SettingsGroup:
    """ Some Dummy class for grouping settings """

    def __str__(self):
        str_eq = ''
        for elem in self.__dict__:
            str_eq = str_eq + elem + ': ' + str(self.__dict__[elem]) + '\n'
        return str_eq

    def __repr__(self):
        return '\n' + str(self.__str__()) + '\n'


if __name__ == '__main__':
    os.system("cls")
    print("----- Starte Bierhelper Tool ----- \n")
    if not os.path.exists(SETTINGS_FILE):
        if ask_user_yn("No settings file found, generate default one?"):
            BierListeTool.generate_default_settingfile()
    if not os.path.exists(EXAMPLE_EXCEL):
        if os.path.exists(resource_path(EXAMPLE_EXCEL)):
            import shutil
            shutil.copy2(resource_path(EXAMPLE_EXCEL), os.path.join(os.getcwd(), EXAMPLE_EXCEL))
            logging.static_debug("Copied example excel file from container to directory")
    if not os.path.exists(EXAMPLE_EXCEL) or not os.path.exists(SETTINGS_FILE):
        raise FileNotFoundError("Settings file '{}' or Example file '{}' not found!".format(SETTINGS_FILE, EXAMPLE_EXCEL))

    tool = BierListeTool()
    print("\t ... Done")
